from fastapi import APIRouter, Depends, Query, HTTPException, Response
from fastapi.responses import JSONResponse, FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from slugify import slugify
from datetime import date
import json
import csv
import io
import os
from typing import Optional, List
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, timedelta
from databases.database import get_db
from services.addreport_services import get_filtered_data, get_date_range_from_type
from services.email_util import send_email_report
from models.chart_model import ChartType
from models.category_model import Category
from models.sub_category_model import SubCategory    
from models.product_model import Product
from models.manufacturer_model import Manufacturer
from models.crawl_websites import CrawlingWebsiteLog
from models.crawl_websites_model import CrawlingWebsite
from schemas.crawl_schema import WebsiteCrawlSummary
from schemas.addreport_schema import SaveReportView
from schemas.addreport_schema import ReportSummary
from schemas.addreport_schema import ReportTypeOut
from models.cms_country import CMSCountry
from models.cms_accountcategory import CMSAccountsCategoryLinks
from models.cms_category import CMSCategory
from models.cms_subcategory import CMSSubCategory
from models.cms_brand import CMSManufacturerBrand
from models.cms_summaryproduct import CMSSummaryProduct
from models.cms_model import CMSSummaryProductModel
from models.crawl_websites_model import CrawlingWebsite
from models.cms_price import CMSCrawlingWebsiteProduct
from models.cms_pricehistory import CMSProductPriceHistory
from models.cms_crawlhistory import CMSCrawlingWebsiteProductsCrawlHistory
from models.cms_clientreports import CMSClientReport
from models.cms_cache import CMSCache
from models.cms_parameters import CMSMSTParameter
from models.cms_clientreportparameter import CMSClientReportParameter
from models.cms_description_split import CMSSummaryDescriptionSplit
from models.cms_description_split_desc_link import CMSSummaryDescriptionSplitDescriptionLink
from models.cms_description_split_product_link import CMSSummaryDescriptionSplitProductLink
from models.cms_reporttypes import MSTClientReportType
from models.cms_city import City
from models.cms_publications import CMSNewspaper
from models.cms_retailername import CMSRetailer
from security.security import get_current_user_id


router = APIRouter()

# @router.get("/report-data")
# def get_report_data_all(
#     report_name: str = Query(None),
#     chart_type: str = Query(None),
#     category: str = Query(None),
#     subcategory: str = Query(None),
#     product: str = Query(None),
#     manufacturer: str = Query(None),
#     db: Session = Depends(get_db)
# ):
#     filters = {
#         "report_name": report_name,
#         "chart_type_id": chart_type,
#         "category": category,
#         "subcategory": subcategory,
#         "product": product,
#         "manufacturer": manufacturer
#     }
#     data = get_filtered_data(db, filters)
#     return {"data": data}

# #  Dropdown endpoints
# @router.get("/chart-types")
# def get_chart_types_dropdown(db: Session = Depends(get_db)):
#     return db.query(ChartType).all()

# @router.get("/categories")
# def get_categories_dropdown(db: Session = Depends(get_db)):
#     return db.query(Category).all()

# @router.get("/subcategories")
# def get_subcategories_dropdown(category_id: int, db: Session = Depends(get_db)):
#     return db.query(SubCategory).filter(SubCategory.category_id == category_id).all()

# @router.get("/products")
# def get_products_dropdown(subcategory_id: int, db: Session = Depends(get_db)):
#     return db.query(Product).filter(Product.subcategory_id == subcategory_id).all()

# @router.get("/manufacturers")
# def get_manufacturers_dropdown(product_id: int, db: Session = Depends(get_db)):
#     return db.query(Manufacturer).filter(Manufacturer.product_id == product_id).all()

# crawl website data 
@router.get("/website-crawled-summary", response_model=list[WebsiteCrawlSummary])
def get_today_website_crawled_summary(db: Session = Depends(get_db)):
    today_date = int(datetime.now().strftime("%Y%m%d"))

    results = (
        db.query(
            CrawlingWebsite.companyName.label("websiteName"),
            func.sum(CrawlingWebsiteLog.productsCrawled).label("totalProductsCrawled")
        )
        .join(CrawlingWebsiteLog, CrawlingWebsite.id == CrawlingWebsiteLog.websiteId)
        .filter(
            CrawlingWebsiteLog.dateCrawl == today_date,
            CrawlingWebsiteLog.deleted == 0,
            CrawlingWebsite.deleted == 0
        )
        .group_by(CrawlingWebsiteLog.websiteId)
        .order_by(func.sum(CrawlingWebsiteLog.productsCrawled).desc())
        .all()
    )

    return results

# Report Types
@router.get("/dropdown/report-types", response_model=List[ReportTypeOut])
def get_report_types(db: Session = Depends(get_db)):
    return db.query(MSTClientReportType).all()

# Country Dropdown
@router.get("/dropdown/countries")
def get_countries(db: Session = Depends(get_db)):
    countries = db.query(CMSCountry).all()
    return [{"id": country.country_code, "name": country.country} for country in countries]

# Category Dropdown
@router.get("/dropdown/categories")
def get_categories(country_id: int, db: Session = Depends(get_db)):  
    # Step 1: Fetch the country from CMSCountry table using country_code
    country = db.query(CMSCountry).filter(CMSCountry.country_code == country_id).first()
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")

    # Step 2: Get categoryIds linked to this country from the accounts_category_links table
    category_links = db.query(CMSAccountsCategoryLinks)\
        .filter(CMSAccountsCategoryLinks.countryId == country_id)\
        .all()

    # Step 3: Fetch categories by categoryId from CMSCategory table
    category_ids = [link.categoryId for link in category_links]
    categories = db.query(CMSCategory).filter(CMSCategory.iCategoryCode.in_(category_ids)).all()

    # Return categories as string values (name)
    return [{"id": category.iCategoryCode, "name": category.sCategoryName} for category in categories]

# Subcategory Dropdown
@router.get("/dropdown/subcategories")
def get_subcategories(category_id: int, db: Session = Depends(get_db)):
    subcategories = db.query(CMSSubCategory).filter(CMSSubCategory.iCategoryCode == category_id).all()
    return [{"id": subcategory.iCatSubCode, "name": subcategory.sCatSubName} for subcategory in subcategories]

# Brand Dropdown (Corrected based on ManufacturerBrand and SummaryProduct)
@router.get("/dropdown/brands")
def get_brands(category_id: int, subcategory_id: int, db: Session = Depends(get_db)):
    # Fetching brands from ManufacturerBrand table, linked to category and subcategory via SummaryProduct
    brands = db.query(CMSManufacturerBrand).join(CMSSummaryProduct, CMSSummaryProduct.iProductBrandCode == CMSManufacturerBrand.iManBrandCode)\
                                         .filter(CMSSummaryProduct.iProductCategoryCode == category_id)\
                                         .filter(CMSSummaryProduct.iProductCatSubCode == subcategory_id)\
                                         .distinct().all()

    return [{"id": brand.iManBrandCode, "name": brand.sBrandName} for brand in brands]

# Model Dropdown type
@router.get("/dropdown/models")
def get_models(category_id: int, subcategory_id: int, brand_id: int, db: Session = Depends(get_db)):
    models = db.query(CMSSummaryProductModel).filter(CMSSummaryProductModel.iProductCategoryCode == category_id)\
                                            .filter(CMSSummaryProductModel.iProductCatSubCode == subcategory_id)\
                                            .filter(CMSSummaryProductModel.iProductBrandCode == brand_id).all()
    return [{"id": model.id, "model_no": model.sProductModelNo} for model in models]
# Model Dropdown field
@router.get("/dropdown/descriptions")
def get_descriptions(category_id: int, country_id: int, db: Session = Depends(get_db)):
    descriptions = db.query(CMSSummaryDescriptionSplit).filter(
        CMSSummaryDescriptionSplit.iCategoryCode == category_id,
        CMSSummaryDescriptionSplit.iCountryCode == country_id,
        CMSSummaryDescriptionSplit.deleted == 0
    ).all()

    return [{"id": desc.id, "name": desc.sDescriptionHeading} for desc in descriptions]

# Description Field Dropdown based on selected Description
@router.get("/dropdown/description-fields")
def get_description_fields(description_id: int, db: Session = Depends(get_db)):
    description_fields = db.query(CMSSummaryDescriptionSplitDescriptionLink).filter(
        CMSSummaryDescriptionSplitDescriptionLink.summaryDescriptionSplitId == description_id,
        CMSSummaryDescriptionSplitDescriptionLink.deleted == 0
    ).all()

    return [{"id": field.id, "name": field.descriptionField} for field in description_fields]

 
# Retailer Dropdown
@router.get("/dropdown/retailers")
def get_retailers(country_id: int, db: Session = Depends(get_db)):
    # Get retailers (company names) from CrawlingWebsite based on country code
    retailers = db.query(CrawlingWebsite).filter(
        CrawlingWebsite.countryCode == country_id
    ).all()

    return [{"id": r.id, "name": r.companyName} for r in retailers]

# City Dropdown
@router.get("/dropdown/cities")
def get_cities(country_code: int, db: Session = Depends(get_db)):
    # Join tblCity and tblRetailer to return only cities with active retailers
    cities = db.query(City).join(
        CMSRetailer,
        City.iCityCode == CMSRetailer.iRetailerCityCode
    ).filter(
        City.iCountryCode == country_code,
        City.bDeleted != 1,
        City.bActivated == 1,
        CMSRetailer.bDeleted != 1
    ).distinct().order_by(City.sCityName).all()

    return [{"id": city.iCityCode, "name": city.sCityName} for city in cities]

# Publication Dropdown
@router.get("/dropdown/publications")
def get_publications(city_code: int, db: Session = Depends(get_db)):
    # Filter publications based on city and non-deleted entries
    publications = db.query(CMSNewspaper).filter(
        CMSNewspaper.iNewspaperCityCode == city_code,
        CMSNewspaper.bDeleted != 1
    ).distinct().order_by(CMSNewspaper.sNewspaperName).all()

    return [
        {
            "id": pub.iNewspaperCode,
            "name": pub.sNewspaperName,
            "publication": pub.sNewspaperPublication
        }
        for pub in publications
    ]

# Submit endpoint
@router.get("/submit")
def get_price_by_names(
    country: str = Query(...),
    category: str = Query(...),
    subcategory: str = Query(...),
    brand: str = Query(...),
    model: str = Query(...),
    retailer: str = Query(...),
    description_type: Optional[str] = Query(None, description="Choose 'Name' or 'Features'"),
    description_field: Optional[str] = Query(None, description="Actual Description Field value for fetching product IDs"),
    start_date: str = Query(None, description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(None, description="End date in YYYY-MM-DD format"),
    date_type: Optional[str] = Query(None, description="Optional date type: MTD, QTD, Q1, Q2, etc."),
    db: Session = Depends(get_db)
):
    # Step 1: Resolve names to codes
    country_obj = db.query(CMSCountry).filter(CMSCountry.country == country).first()
    category_obj = db.query(CMSCategory).filter(CMSCategory.sCategoryName == category).first()
    subcategory_obj = db.query(CMSSubCategory).filter(CMSSubCategory.sCatSubName == subcategory).first()
    brand_obj = db.query(CMSManufacturerBrand).filter(CMSManufacturerBrand.sBrandName == brand).first()
    retailer_obj = db.query(CrawlingWebsite).filter(CrawlingWebsite.companyName == retailer).first()

    if not all([country_obj, category_obj, subcategory_obj, brand_obj, retailer_obj]):
        raise HTTPException(status_code=404, detail="One or more values not found")

    # Step 2: Match summary_products using selected filters
    summary_product = (
        db.query(CMSSummaryProduct)
        .filter(CMSSummaryProduct.iProductCountryCode == country_obj.country_code)
        .filter(CMSSummaryProduct.iProductCategoryCode == category_obj.iCategoryCode)
        .filter(CMSSummaryProduct.iProductCatSubCode == subcategory_obj.iCatSubCode)
        .filter(CMSSummaryProduct.sProductModelNo == model)
        .filter(CMSSummaryProduct.iProductBrandCode == brand_obj.iManBrandCode)
        .first()
    )

    if not summary_product:
        return {
            "country": country,
            "category": category,
            "subcategory": subcategory,
            "brand": brand,
            "model": model,
            "retailer": retailer,
            "message": "Matching product not found in summary_products"
        }
    
    final_product_ids = []

    if description_type and description_field:
        # Step 3: Handle description filtering
        if description_type not in ["Name", "Features"]:
            raise HTTPException(status_code=400, detail="description_type must be either 'Name' or 'Features'")

        # Find split name ID for description
        split_name_obj = db.query(CMSSummaryDescriptionSplit).filter(CMSSummaryDescriptionSplit.sDescriptionHeading == description_type).first()
        if not split_name_obj:
            raise HTTPException(status_code=404, detail="Provided description field not found")

        # Find description link
        split_link_obj = (
            db.query(CMSSummaryDescriptionSplitDescriptionLink)
            .filter(CMSSummaryDescriptionSplitDescriptionLink.summaryDescriptionSplitId == split_name_obj.id)
            .first()
        )

        if not split_link_obj:
            raise HTTPException(status_code=404, detail="Description link not found")

        # Find mapped products
        mapped_products = (
            db.query(CMSSummaryDescriptionSplitProductLink)
            .filter(CMSSummaryDescriptionSplitProductLink.summaryDescriptionSplitId == split_link_obj.summaryDescriptionSplitId)
            .filter(CMSSummaryDescriptionSplitProductLink.summaryDescriptionSplitDescriptionId == split_link_obj.id)
            .filter(CMSSummaryDescriptionSplitProductLink.deleted == 0)
            .all()
        )

        if not mapped_products:
            raise HTTPException(status_code=404, detail="No mapped products found")

        mapped_product_codes = [mp.iProductCode for mp in mapped_products]

        # Fetch mapped products from summary_products
        final_summary_products = (
            db.query(CMSSummaryProduct)
            .filter(CMSSummaryProduct.iProductCode.in_(mapped_product_codes))
            .filter(CMSSummaryProduct.deleted == 0)
            .all()
        )
        final_product_codes = [fsp.iProductCode for fsp in final_summary_products]

        # Fetch corresponding crawling website products
        crawling_products = (
            db.query(CMSCrawlingWebsiteProduct)
            .filter(CMSCrawlingWebsiteProduct.systemProductId.in_(final_product_codes))
            .filter(CMSCrawlingWebsiteProduct.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProduct.isActive == 1)
            .all()
        )

    else:
        # Step 3: Use iProductCode to get products
        crawling_products = (
            db.query(CMSCrawlingWebsiteProduct)
            .filter(CMSCrawlingWebsiteProduct.systemProductId == summary_product.iProductCode)
            .filter(CMSCrawlingWebsiteProduct.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProduct.isActive == 1)
            .all()
        )

        if not crawling_products:
            return {
                "country": country,
                "category": category,
                "subcategory": subcategory,
                "brand": brand,
                "model": model,
                "retailer": retailer,
                "message": "Product found in summary, but not in crawling website products"
            }
    
    # Special Case: If date_type is "current_day", return current price
    if date_type == "current_day":
        today_int = int(date.today().strftime("%Y%m%d"))
        crawl_histories = (
            db.query(CMSCrawlingWebsiteProductsCrawlHistory)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlDate == today_int)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.deleted == 0)
            .all()
        )
        valid_product_ids = set()

        for history in crawl_histories:
            serialized = history.searializeProductCodes
            for product in crawling_products:
                pid = product.id
                if f'i:{pid};' in serialized or f's:{len(str(pid))}:"{pid}"' in serialized:
                    valid_product_ids.add(pid)
        
        # Filter crawling_products to include only those in crawl history
        filtered_products = [p for p in crawling_products if p.id in valid_product_ids]

        if not filtered_products:
           return {
                "country": country,
                "category": category,
                "subcategory": subcategory,
                "brand": brand,
                "model": model,
                "retailer": retailer,
                "message": "No valid product found in crawl history for current day"
            }
        result = []
        for product in filtered_products:
            description_field = ""

            if description_type == "Name":
                description_field = product.description
            elif description_type == "Features":
                summary_product_obj = db.query(CMSSummaryProduct).filter(
                    CMSSummaryProduct.iProductCode == product.systemProductId
                ).first()
                description_field = summary_product_obj.sProductOneLineDesc if summary_product_obj else ""
        
            entry = {
                    "country": country,
                    "category": category,
                    "subcategory": subcategory,
                    "brand": brand,
                    "model": model,
                    "retailer": retailer,
                    "price": f"${product.price:,.2f}" if product.price else "N/A",
            }

            # Add description_type as dynamic key
            if description_type and description_field:
                entry[description_type] = description_field

            result.append(entry)

        return result
        
    # Step 4: Handle date filters
    if start_date and end_date:
       try:
           start_dt = datetime.strptime(start_date, "%Y-%m-%d")
           end_dt = datetime.strptime(end_date, "%Y-%m-%d")
           start = int(start_dt.strftime("%Y%m%d"))
           end = int(end_dt.strftime("%Y%m%d"))
       except ValueError:
             raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    elif date_type:
        start_dt, end_dt = get_date_range_from_type(date_type)
        if not start_dt or not end_dt:
           raise HTTPException(status_code=400, detail="Invalid date_type value")
        start = int(start_dt.strftime("%Y%m%d"))
        end = int(end_dt.strftime("%Y%m%d"))
    else:
        raise HTTPException(status_code=400, detail="Please provide a date range or date_type to fetch price history")

    # Step 5: Match product IDs with crawl history via serialized codes
    crawl_histories = (
        db.query(CMSCrawlingWebsiteProductsCrawlHistory)
        .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSCrawlingWebsiteProductsCrawlHistory.deleted == 0)
        .all()
    )

    matched_entries = defaultdict(list)

    # Group crawl records by crawlDate where a match with product ID exists
    for history in crawl_histories:
        serialized = history.searializeProductCodes
        for product in crawling_products:
            pid = product.id
            if f'i:{pid};' in serialized or f's:{len(str(pid))}:"{pid}"' in serialized:
                if start <= history.crawlDate <= end:
                    matched_entries[history.crawlDate].append({
                        "product_id": pid,
                        "crawl_date": history.crawlDate,
                        "created_at": history.created
                    })
    
    # Now you need to process each crawlDate and select the latest record for each one
    final_matched_entries = []

    # For each crawlDate, pick the record with latest created_at
    for crawl_date, records in matched_entries.items():
         latest_record = max(records, key=lambda x: x["created_at"])

         # Add the latest record to final list
         final_matched_entries.append({
             "product_id": latest_record["product_id"],
             "crawl_date": latest_record["crawl_date"]
        })

    if not matched_entries:
        return {
            "country": country,
            "category": category,
            "subcategory": subcategory,
            "brand": brand,
            "model": model,
            "retailer": retailer,
            "message": "No matching crawl history data found for the given products and date range"
        }

    # Step 6: Get all matched product IDs and latest created crawlDate per product
    product_latest_crawl = {}

    for record in final_matched_entries:
        pid = record["product_id"]
        crawl_date = record["crawl_date"]
        if pid not in product_latest_crawl or product_latest_crawl[pid]["created_at"] < record.get("created_at", datetime.min):
            product_latest_crawl[pid] = {
                "crawl_date": crawl_date,
                "created_at": record.get("created_at", datetime.min)
            }

    # Step 7: Get latest price history reference date per product
    latest_price_history = (
        db.query(CMSProductPriceHistory.productId, func.max(CMSProductPriceHistory.referenceDate).label("last_ref"))
        .filter(CMSProductPriceHistory.productId.in_(product_latest_crawl.keys()))
        .filter(CMSProductPriceHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSProductPriceHistory.deleted == 0)
        .group_by(CMSProductPriceHistory.productId)
        .all()
    )

    product_last_ref_date = {
        p.productId: p.last_ref for p in latest_price_history
    }

    # Load full history for all matched products
    full_history_prices = defaultdict(dict)

    history_records = (
        db.query(CMSProductPriceHistory)
        .filter(CMSProductPriceHistory.productId.in_(product_latest_crawl.keys()))
        .filter(CMSProductPriceHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSProductPriceHistory.deleted == 0)
        .all()
    )

    for record in history_records:
        full_history_prices[record.productId][record.referenceDate] = record.price


    # Step 8: Get crawling website product prices (used for forward fill)
    product_prices = {
        p.id: p.price for p in crawling_products if p.price is not None
    }

    # Step 9: Build complete timeline
    final_result = []
    all_dates = []
    cursor = start_dt.date()
    while cursor <= end_dt.date():
        all_dates.append(cursor)
        cursor += timedelta(days=1)
    
    today_date = date.today()
    for pid in product_latest_crawl.keys():
        crawl_date = product_latest_crawl[pid]["crawl_date"]
        forward_price = product_prices.get(pid)
        history_map = full_history_prices.get(pid, {})

        # Sort the historical dates ascending for forward lookup
        sorted_history_dates = sorted(history_map.keys())
        price_timeline = {}
        last_seen_price = None

        # Step 1: Populate known historical prices
        for ref_date in sorted_history_dates:
            price_timeline[ref_date] = history_map[ref_date]

        # Step 2: Backfill missing dates with the last known price
        current_idx = len(sorted_history_dates) - 1
        for date_obj in reversed(all_dates):
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int in price_timeline:
                last_seen_price = price_timeline[date_int]   
            elif last_seen_price is not None and date_int < sorted_history_dates[-1]:
                price_timeline[date_int] = last_seen_price

        # Step 3: Forward-fill using crawling website price
        for date_obj in all_dates:
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int > int(today_date.strftime("%Y%m%d")):
                continue

            if date_int >= crawl_date and date_int not in price_timeline and forward_price is not None:
                price_timeline[date_int] = forward_price

        # Step 4: Format results
        result = []
        for date_obj in all_dates:
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int in price_timeline:
                final_description = ""
        
                if description_type == "Name":
                    crawling_product = next((p for p in crawling_products if p.id == pid), None)
                    final_description = crawling_product.description if crawling_product else ""

                elif description_type == "Features":
                    summary_product_obj = db.query(CMSSummaryProduct).filter(
                        CMSSummaryProduct.iProductCode == crawling_product.systemProductId
                    ).first() if crawling_product else None
                    final_description = summary_product_obj.sProductOneLineDesc if summary_product_obj else ""

            entry = {
                    "country": country,
                    "category": category,
                    "subcategory": subcategory,
                    "brand": brand,
                    "model": model,
                    "retailer": retailer,
                    "Date": date_obj.strftime("%d-%m-%Y"),
                    "price": f"${price_timeline[date_int]:,.2f}"
                }

            # Dynamically insert the description type as a key
            if description_type and final_description:
                entry[description_type] = final_description

            result.append(entry)

        return result

@router.get("/export-excel")
def get_price_by_names(
    country: str = Query(...),
    category: str = Query(...),
    subcategory: str = Query(...),
    brand: str = Query(...),
    model: str = Query(...),
    retailer: str = Query(...),
    description_type: Optional[str] = Query(None, description="Choose 'Name' or 'Features'"),
    description_field: Optional[str] = Query(None, description="Actual Description Field value for fetching product IDs"),
    start_date: str = Query(None, description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(None, description="End date in YYYY-MM-DD format"),
    date_type: Optional[str] = Query(None, description="Optional date type: MTD, QTD, Q1, Q2, etc."),
    db: Session = Depends(get_db)
):
    # Step 1: Resolve names to codes
    country_obj = db.query(CMSCountry).filter(CMSCountry.country == country).first()
    category_obj = db.query(CMSCategory).filter(CMSCategory.sCategoryName == category).first()
    subcategory_obj = db.query(CMSSubCategory).filter(CMSSubCategory.sCatSubName == subcategory).first()
    brand_obj = db.query(CMSManufacturerBrand).filter(CMSManufacturerBrand.sBrandName == brand).first()
    retailer_obj = db.query(CrawlingWebsite).filter(CrawlingWebsite.companyName == retailer).first()

    if not all([country_obj, category_obj, subcategory_obj, brand_obj, retailer_obj]):
        raise HTTPException(status_code=404, detail="One or more values not found")

    # Step 2: Match summary_products using selected filters
    summary_product = (
        db.query(CMSSummaryProduct)
        .filter(CMSSummaryProduct.iProductCountryCode == country_obj.country_code)
        .filter(CMSSummaryProduct.iProductCategoryCode == category_obj.iCategoryCode)
        .filter(CMSSummaryProduct.iProductCatSubCode == subcategory_obj.iCatSubCode)
        .filter(CMSSummaryProduct.sProductModelNo == model)
        .filter(CMSSummaryProduct.iProductBrandCode == brand_obj.iManBrandCode)
        .first()
    )

    if not summary_product:
        return {
            "country": country,
            "category": category,
            "subcategory": subcategory,
            "brand": brand,
            "model": model,
            "retailer": retailer,
            "message": "Matching product not found in summary_products"
        }
        final_product_ids = []

    if description_type and description_field:
        # Step 3: Handle description filtering
        if description_type not in ["Name", "Features"]:
            raise HTTPException(status_code=400, detail="description_type must be either 'Name' or 'Features'")

        # Find split name ID for description
        split_name_obj = db.query(CMSSummaryDescriptionSplit).filter(CMSSummaryDescriptionSplit.sDescriptionHeading == description_type).first()
        if not split_name_obj:
            raise HTTPException(status_code=404, detail="Provided description field not found")

        # Find description link
        split_link_obj = (
            db.query(CMSSummaryDescriptionSplitDescriptionLink)
            .filter(CMSSummaryDescriptionSplitDescriptionLink.summaryDescriptionSplitId == split_name_obj.id)
            .first()
        )

        if not split_link_obj:
            raise HTTPException(status_code=404, detail="Description link not found")

        # Find mapped products
        mapped_products = (
            db.query(CMSSummaryDescriptionSplitProductLink)
            .filter(CMSSummaryDescriptionSplitProductLink.summaryDescriptionSplitId == split_link_obj.summaryDescriptionSplitId)
            .filter(CMSSummaryDescriptionSplitProductLink.summaryDescriptionSplitDescriptionId == split_link_obj.id)
            .filter(CMSSummaryDescriptionSplitProductLink.deleted == 0)
            .all()
        )

        if not mapped_products:
            raise HTTPException(status_code=404, detail="No mapped products found")

        mapped_product_codes = [mp.iProductCode for mp in mapped_products]

        # Fetch mapped products from summary_products
        final_summary_products = (
            db.query(CMSSummaryProduct)
            .filter(CMSSummaryProduct.iProductCode.in_(mapped_product_codes))
            .filter(CMSSummaryProduct.deleted == 0)
            .all()
        )
        final_product_codes = [fsp.iProductCode for fsp in final_summary_products]

        # Fetch corresponding crawling website products
        crawling_products = (
            db.query(CMSCrawlingWebsiteProduct)
            .filter(CMSCrawlingWebsiteProduct.systemProductId.in_(final_product_codes))
            .filter(CMSCrawlingWebsiteProduct.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProduct.isActive == 1)
            .all()
        )

    else:
        # Step 3: Use iProductCode to get products
        crawling_products = (
            db.query(CMSCrawlingWebsiteProduct)
            .filter(CMSCrawlingWebsiteProduct.systemProductId == summary_product.iProductCode)
            .filter(CMSCrawlingWebsiteProduct.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProduct.isActive == 1)
            .all()
        )

        if not crawling_products:
            return {
                "country": country,
                "category": category,
                "subcategory": subcategory,
                "brand": brand,
                "model": model,
                "retailer": retailer,
                "message": "Product found in summary, but not in crawling website products"
            }
    
    # Special Case: If date_type is "current_day", return current price
    if date_type == "current_day":
        today_int = int(date.today().strftime("%Y%m%d"))
        crawl_histories = (
            db.query(CMSCrawlingWebsiteProductsCrawlHistory)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlDate == today_int)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.deleted == 0)
            .all()
        )
        valid_product_ids = set()

        for history in crawl_histories:
            serialized = history.searializeProductCodes
            for product in crawling_products:
                pid = product.id
                if f'i:{pid};' in serialized or f's:{len(str(pid))}:"{pid}"' in serialized:
                    valid_product_ids.add(pid)
        
        # Filter crawling_products to include only those in crawl history
        filtered_products = [p for p in crawling_products if p.id in valid_product_ids]

        if not filtered_products:
           return {
                "country": country,
                "category": category,
                "subcategory": subcategory,
                "brand": brand,
                "model": model,
                "retailer": retailer,
                "message": "No valid product found in crawl history for current day"
            }
        result = []
        for product in filtered_products:
            description_field = ""

            if description_type == "Name":
                description_field = product.description
            elif description_type == "Features":
                summary_product_obj = db.query(CMSSummaryProduct).filter(
                    CMSSummaryProduct.iProductCode == product.systemProductId
                ).first()
                description_field = summary_product_obj.sProductOneLineDesc if summary_product_obj else ""
        
            result.append({
                "country": country,
                "category": category,
                "subcategory": subcategory,
                "brand": brand,
                "model": model,
                "retailer": retailer,
                "description_type": description_type or "",
                "description_field": description_field or "",
                "price": f"${product.price:,.2f}" if product.price else "N/A",
            })

        return result
                 
    # Step 4: Handle date filters
    if start_date and end_date:
       try:
           start_dt = datetime.strptime(start_date, "%Y-%m-%d")
           end_dt = datetime.strptime(end_date, "%Y-%m-%d")
           start = int(start_dt.strftime("%Y%m%d"))
           end = int(end_dt.strftime("%Y%m%d"))
       except ValueError:
             raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    elif date_type:
        start_dt, end_dt = get_date_range_from_type(date_type)
        if not start_dt or not end_dt:
           raise HTTPException(status_code=400, detail="Invalid date_type value")
        start = int(start_dt.strftime("%Y%m%d"))
        end = int(end_dt.strftime("%Y%m%d"))
    else:
        raise HTTPException(status_code=400, detail="Please provide a date range or date_type to fetch price history")

    # Step 5: Match product IDs with crawl history via serialized codes
    crawl_histories = (
        db.query(CMSCrawlingWebsiteProductsCrawlHistory)
        .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSCrawlingWebsiteProductsCrawlHistory.deleted == 0)
        .all()
    )

    matched_entries = defaultdict(list)

    # Group crawl records by crawlDate where a match with product ID exists
    for history in crawl_histories:
        serialized = history.searializeProductCodes
        for product in crawling_products:
            pid = product.id
            if f'i:{pid};' in serialized or f's:{len(str(pid))}:"{pid}"' in serialized:
                if start <= history.crawlDate <= end:
                    matched_entries[history.crawlDate].append({
                        "product_id": pid,
                        "crawl_date": history.crawlDate,
                        "created_at": history.created
                    })
    
    # Now you need to process each crawlDate and select the latest record for each one
    final_matched_entries = []

    # For each crawlDate, pick the record with latest created_at
    for crawl_date, records in matched_entries.items():
         latest_record = max(records, key=lambda x: x["created_at"])

         # Add the latest record to final list
         final_matched_entries.append({
             "product_id": latest_record["product_id"],
             "crawl_date": latest_record["crawl_date"]
        })

    if not matched_entries:
        return {
            "country": country,
            "category": category,
            "subcategory": subcategory,
            "brand": brand,
            "model": model,
            "retailer": retailer,
            "message": "No matching crawl history data found for the given products and date range"
        }

    # Step 6: Get all matched product IDs and latest created crawlDate per product
    product_latest_crawl = {}

    for record in final_matched_entries:
        pid = record["product_id"]
        crawl_date = record["crawl_date"]
        if pid not in product_latest_crawl or product_latest_crawl[pid]["created_at"] < record.get("created_at", datetime.min):
            product_latest_crawl[pid] = {
                "crawl_date": crawl_date,
                "created_at": record.get("created_at", datetime.min)
            }

    # Step 7: Get latest price history reference date per product
    latest_price_history = (
        db.query(CMSProductPriceHistory.productId, func.max(CMSProductPriceHistory.referenceDate).label("last_ref"))
        .filter(CMSProductPriceHistory.productId.in_(product_latest_crawl.keys()))
        .filter(CMSProductPriceHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSProductPriceHistory.deleted == 0)
        .group_by(CMSProductPriceHistory.productId)
        .all()
    )

    product_last_ref_date = {
        p.productId: p.last_ref for p in latest_price_history
    }

    # Load full history for all matched products
    full_history_prices = defaultdict(dict)

    history_records = (
        db.query(CMSProductPriceHistory)
        .filter(CMSProductPriceHistory.productId.in_(product_latest_crawl.keys()))
        .filter(CMSProductPriceHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSProductPriceHistory.deleted == 0)
        .all()
    )

    for record in history_records:
        full_history_prices[record.productId][record.referenceDate] = record.price


    # Step 8: Get crawling website product prices (used for forward fill)
    product_prices = {
        p.id: p.price for p in crawling_products if p.price is not None
    }

    # Step 9: Build complete timeline
    final_result = []
    all_dates = []
    cursor = start_dt.date()
    while cursor <= end_dt.date():
        all_dates.append(cursor)
        cursor += timedelta(days=1)
    
    today_date = date.today()
    for pid in product_latest_crawl.keys():
        crawl_date = product_latest_crawl[pid]["crawl_date"]
        forward_price = product_prices.get(pid)
        history_map = full_history_prices.get(pid, {})

        # Sort the historical dates ascending for forward lookup
        sorted_history_dates = sorted(history_map.keys())
        price_timeline = {}
        last_seen_price = None

        # Step 1: Populate known historical prices
        for ref_date in sorted_history_dates:
            price_timeline[ref_date] = history_map[ref_date]

        # Step 2: Backfill missing dates with the last known price
        current_idx = len(sorted_history_dates) - 1
        for date_obj in reversed(all_dates):
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int in price_timeline:
                last_seen_price = price_timeline[date_int]   
            elif last_seen_price is not None and date_int < sorted_history_dates[-1]:
                price_timeline[date_int] = last_seen_price

        # Step 3: Forward-fill using crawling website price
        for date_obj in all_dates:
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int > int(today_date.strftime("%Y%m%d")):
                continue

            if date_int >= crawl_date and date_int not in price_timeline and forward_price is not None:
                price_timeline[date_int] = forward_price

        # Step 4: Format results
        for date_obj in all_dates:
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int in price_timeline:
                final_description = ""
        
                if description_type == "Name":
                    crawling_product = next((p for p in crawling_products if p.id == pid), None)
                    final_description = crawling_product.description if crawling_product else ""

                elif description_type == "Features":
                    summary_product_obj = db.query(CMSSummaryProduct).filter(
                        CMSSummaryProduct.iProductCode == crawling_product.systemProductId
                    ).first() if crawling_product else None
                    final_description = summary_product_obj.sProductOneLineDesc if summary_product_obj else ""

            entry = {
                    "country": country,
                    "category": category,
                    "subcategory": subcategory,
                    "brand": brand,
                    "model": model,
                    "retailer": retailer,
                    "Date": date_obj.strftime("%d-%m-%Y"),
                    "price": f"${price_timeline[date_int]:,.2f}"
            }

            # Dynamically insert the description type as a key
            if description_type and final_description:
                entry[description_type] = final_description

            final_result.append(entry)
                
    start_dt = end_dt = None

    if date_type and not (start_date and end_date):
        start_dt, end_dt = get_date_range_from_type(date_type)
        if not start_dt or not end_dt:
            raise HTTPException(status_code=400, detail="Invalid date_type provided.")
        start_date = start_dt.strftime("%Y-%m-%d")
        end_date = end_dt.strftime("%Y-%m-%d")

    elif date_type and (start_date or end_date):
        raise HTTPException(status_code=400, detail="Provide either date_type OR start_date/end_date — not both.")

    if start_date and end_date:
        try:
            if not start_dt:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            if not end_dt:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            
            readable_label = {
                "QTD": "Quarter to Date",
                "MTD": "Month to Date",
                "YTD": "Year to Date",
                "last_month": "Last Month",
                "last_week": "Last Week",
                "current_day": "Today",
                "Q1": "Q1",
                "Q2": "Q2",
                "Q3": "Q3",
                "Q4": "Q4"
            }.get(date_type.lower() if date_type else "", "Date Range")

            report_title = f"{readable_label}: {start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"
        except Exception:
            report_title = f"{date_type or 'Date Range'}: {start_date} to {end_date}"
    else:
        report_title = "Current Product Prices"
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Report"

    # Add title row in merged cells
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"] = report_title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = list(final_result[0].keys()) if final_result else []
    ws.append(headers)

    # Define header style
    header_font = Font(bold=True, color="000000")  # black text
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")  # Blue background
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Apply styles to header row
    for col_num, cell in enumerate(ws[2], 1):  # Row 2 if the title is on row 1
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Optional styling
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in final_result:
        ws.append([row.get(h, "") for h in headers])
    
    file_name = f"price_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join("/tmp", file_name)
    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@router.get("/export-csv")
def export_csv_with_full_logic(
    country: str = Query(...),
    category: str = Query(...),
    subcategory: str = Query(...),
    brand: str = Query(...),
    model: str = Query(...),
    retailer: str = Query(...),
    start_date: str = Query(None),
    end_date: str = Query(None),
    date_type: str = Query(None),
    db: Session = Depends(get_db)
):
    # Step 1: Reuse shared function
    response = get_price_by_names(
        db=db,
        country=country,
        category=category,
        subcategory=subcategory,
        brand=brand,
        model=model,
        retailer=retailer,
        start_date=start_date,
        end_date=end_date,
        date_type=date_type
    )

    # Extract JSON data from the response content
    content_str = response.body.decode() if hasattr(response, 'body') else response.body
    data_rows = json.loads(content_str)

    # Step 2: Prepare CSV
    output = io.StringIO()
    fieldnames = [
        "Date", "Country", "Category", "Subcategory", "Brand", "Model",
        "Retailer", "Price"
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for row in data_rows:
        writer.writerow(row)

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={brand}_Report.csv"}
    )

@router.get("/export-pdf")
def get_price_by_names(
    country: str = Query(...),
    category: str = Query(...),
    subcategory: str = Query(...),
    brand: str = Query(...),
    model: str = Query(...),
    retailer: str = Query(...),
    start_date: str = Query(None, description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(None, description="End date in YYYY-MM-DD format"),
    date_type: Optional[str] = Query(None, description="Optional date type: MTD, QTD, Q1, Q2, etc."),
    db: Session = Depends(get_db)
):
    # Step 1: Resolve names to codes
    country_obj = db.query(CMSCountry).filter(CMSCountry.country == country).first()
    category_obj = db.query(CMSCategory).filter(CMSCategory.sCategoryName == category).first()
    subcategory_obj = db.query(CMSSubCategory).filter(CMSSubCategory.sCatSubName == subcategory).first()
    brand_obj = db.query(CMSManufacturerBrand).filter(CMSManufacturerBrand.sBrandName == brand).first()
    retailer_obj = db.query(CrawlingWebsite).filter(CrawlingWebsite.companyName == retailer).first()

    if not all([country_obj, category_obj, subcategory_obj, brand_obj, retailer_obj]):
        raise HTTPException(status_code=404, detail="One or more values not found")

    # Step 2: Match summary_products using selected filters
    summary_product = (
        db.query(CMSSummaryProduct)
        .filter(CMSSummaryProduct.iProductCountryCode == country_obj.country_code)
        .filter(CMSSummaryProduct.iProductCategoryCode == category_obj.iCategoryCode)
        .filter(CMSSummaryProduct.iProductCatSubCode == subcategory_obj.iCatSubCode)
        .filter(CMSSummaryProduct.sProductModelNo == model)
        .filter(CMSSummaryProduct.iProductBrandCode == brand_obj.iManBrandCode)
        .first()
    )

    if not summary_product:
        return {
            "country": country,
            "category": category,
            "subcategory": subcategory,
            "brand": brand,
            "model": model,
            "retailer": retailer,
            "message": "Matching product not found in summary_products"
        }

    else:
        # Step 3: Use iProductCode to get products
        crawling_products = (
            db.query(CMSCrawlingWebsiteProduct)
            .filter(CMSCrawlingWebsiteProduct.systemProductId == summary_product.iProductCode)
            .filter(CMSCrawlingWebsiteProduct.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProduct.isActive == 1)
            .all()
        )

        if not crawling_products:
            return {
                "country": country,
                "category": category,
                "subcategory": subcategory,
                "brand": brand,
                "model": model,
                "retailer": retailer,
                "message": "Product found in summary, but not in crawling website products"
            }
    
    # Special Case: If date_type is "current_day", return current price
    if date_type == "current_day":
        today_int = int(date.today().strftime("%Y%m%d"))
        crawl_histories = (
            db.query(CMSCrawlingWebsiteProductsCrawlHistory)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlWebsiteId == retailer_obj.id)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlDate == today_int)
            .filter(CMSCrawlingWebsiteProductsCrawlHistory.deleted == 0)
            .all()
        )
        valid_product_ids = set()

        for history in crawl_histories:
            serialized = history.searializeProductCodes
            for product in crawling_products:
                pid = product.id
                if f'i:{pid};' in serialized or f's:{len(str(pid))}:"{pid}"' in serialized:
                    valid_product_ids.add(pid)
        
        # Filter crawling_products to include only those in crawl history
        filtered_products = [p for p in crawling_products if p.id in valid_product_ids]

        if not filtered_products:
           return {
                "country": country,
                "category": category,
                "subcategory": subcategory,
                "brand": brand,
                "model": model,
                "retailer": retailer,
                "message": "No valid product found in crawl history for current day"
            }
        return [
           {
            "country": country,
            "category": category,
            "subcategory": subcategory,
            "brand": brand,
            "model": model,
            "retailer": retailer,
            "price": f"${product.price:,.2f}" if product.price else "N/A",
           }

           for product in filtered_products
        ]
        
    # Step 4: Handle date filters
    if start_date and end_date:
       try:
           start_dt = datetime.strptime(start_date, "%Y-%m-%d")
           end_dt = datetime.strptime(end_date, "%Y-%m-%d")
           start = int(start_dt.strftime("%Y%m%d"))
           end = int(end_dt.strftime("%Y%m%d"))
       except ValueError:
             raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    elif date_type:
        start_dt, end_dt = get_date_range_from_type(date_type)
        if not start_dt or not end_dt:
           raise HTTPException(status_code=400, detail="Invalid date_type value")
        start = int(start_dt.strftime("%Y%m%d"))
        end = int(end_dt.strftime("%Y%m%d"))
    else:
        raise HTTPException(status_code=400, detail="Please provide a date range or date_type to fetch price history")

    # Step 5: Match product IDs with crawl history via serialized codes
    crawl_histories = (
        db.query(CMSCrawlingWebsiteProductsCrawlHistory)
        .filter(CMSCrawlingWebsiteProductsCrawlHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSCrawlingWebsiteProductsCrawlHistory.deleted == 0)
        .all()
    )

    matched_entries = defaultdict(list)

    # Group crawl records by crawlDate where a match with product ID exists
    for history in crawl_histories:
        serialized = history.searializeProductCodes
        for product in crawling_products:
            pid = product.id
            if f'i:{pid};' in serialized or f's:{len(str(pid))}:"{pid}"' in serialized:
                if start <= history.crawlDate <= end:
                    matched_entries[history.crawlDate].append({
                        "product_id": pid,
                        "crawl_date": history.crawlDate,
                        "created_at": history.created
                    })
    
    # Now you need to process each crawlDate and select the latest record for each one
    final_matched_entries = []

    # For each crawlDate, pick the record with latest created_at
    for crawl_date, records in matched_entries.items():
         latest_record = max(records, key=lambda x: x["created_at"])

         # Add the latest record to final list
         final_matched_entries.append({
             "product_id": latest_record["product_id"],
             "crawl_date": latest_record["crawl_date"]
        })

    if not matched_entries:
        return {
            "country": country,
            "category": category,
            "subcategory": subcategory,
            "brand": brand,
            "model": model,
            "retailer": retailer,
            "message": "No matching crawl history data found for the given products and date range"
        }

    # Step 6: Get all matched product IDs and latest created crawlDate per product
    product_latest_crawl = {}

    for record in final_matched_entries:
        pid = record["product_id"]
        crawl_date = record["crawl_date"]
        if pid not in product_latest_crawl or product_latest_crawl[pid]["created_at"] < record.get("created_at", datetime.min):
            product_latest_crawl[pid] = {
                "crawl_date": crawl_date,
                "created_at": record.get("created_at", datetime.min)
            }

    # Step 7: Get latest price history reference date per product
    latest_price_history = (
        db.query(CMSProductPriceHistory.productId, func.max(CMSProductPriceHistory.referenceDate).label("last_ref"))
        .filter(CMSProductPriceHistory.productId.in_(product_latest_crawl.keys()))
        .filter(CMSProductPriceHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSProductPriceHistory.deleted == 0)
        .group_by(CMSProductPriceHistory.productId)
        .all()
    )

    product_last_ref_date = {
        p.productId: p.last_ref for p in latest_price_history
    }

    # Load full history for all matched products
    full_history_prices = defaultdict(dict)

    history_records = (
        db.query(CMSProductPriceHistory)
        .filter(CMSProductPriceHistory.productId.in_(product_latest_crawl.keys()))
        .filter(CMSProductPriceHistory.crawlWebsiteId == retailer_obj.id)
        .filter(CMSProductPriceHistory.deleted == 0)
        .all()
    )

    for record in history_records:
        full_history_prices[record.productId][record.referenceDate] = record.price


    # Step 8: Get crawling website product prices (used for forward fill)
    product_prices = {
        p.id: p.price for p in crawling_products if p.price is not None
    }

    # Step 9: Build complete timeline
    final_result = []
    all_dates = []
    cursor = start_dt.date()
    while cursor <= end_dt.date():
        all_dates.append(cursor)
        cursor += timedelta(days=1)
    
    today_date = date.today()
    for pid in product_latest_crawl.keys():
        crawl_date = product_latest_crawl[pid]["crawl_date"]
        forward_price = product_prices.get(pid)
        history_map = full_history_prices.get(pid, {})

        # Sort the historical dates ascending for forward lookup
        sorted_history_dates = sorted(history_map.keys())
        price_timeline = {}
        last_seen_price = None

        # Step 1: Populate known historical prices
        for ref_date in sorted_history_dates:
            price_timeline[ref_date] = history_map[ref_date]

        # Step 2: Backfill missing dates with the last known price
        current_idx = len(sorted_history_dates) - 1
        for date_obj in reversed(all_dates):
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int in price_timeline:
                last_seen_price = price_timeline[date_int]   
            elif last_seen_price is not None and date_int < sorted_history_dates[-1]:
                price_timeline[date_int] = last_seen_price

        # Step 3: Forward-fill using crawling website price
        for date_obj in all_dates:
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int > int(today_date.strftime("%Y%m%d")):
                continue

            if date_int >= crawl_date and date_int not in price_timeline and forward_price is not None:
                price_timeline[date_int] = forward_price

        # Step 4: Format results
        for date_obj in all_dates:
            date_int = int(date_obj.strftime("%Y%m%d"))
            if date_int in price_timeline:
                final_result.append({
                    "country": country,
                    "category": category,
                    "subcategory": subcategory,
                    "brand": brand,
                    "model": model,
                    "retailer": retailer,
                    "Date": date_obj.strftime("%d-%m-%Y"),
                    "price": f"${price_timeline[date_int]:,.2f}"
            })

    # Build table_data from result
    table_data = [list(final_result[0].keys())] if final_result else []
    for row in final_result:
        table_data.append([row.get(col, "") for col in table_data[0]])

    start_dt = end_dt = None

    # 1. Handle date_type logic and convert to actual dates
    if date_type and not (start_date and end_date):
        start_dt, end_dt = get_date_range_from_type(date_type)
        if not start_dt or not end_dt:
            raise HTTPException(status_code=400, detail="Invalid date_type provided.")
        start_date = start_dt.strftime("%Y-%m-%d")
        end_date = end_dt.strftime("%Y-%m-%d")

    elif date_type and (start_date or end_date):
        raise HTTPException(status_code=400, detail="Provide either date_type OR start_date/end_date — not both.")

    # 2. Set the report title based on dates or fallback
    if start_date and end_date:
        try:
            if not start_dt:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            if not end_dt:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            
            readable_label = {
                "QTD": "Quarter to Date",
                "MTD": "Month to Date",
                "YTD": "Year to Date",
                "last_month": "Last Month",
                "last_week": "Last Week",
                "current_day": "Today",
                "Q1": "Q1",
                "Q2": "Q2",
                "Q3": "Q3",
                "Q4": "Q4"
            }.get(date_type.lower() if date_type else "", "Date Range")

            report_title = f"{readable_label}: {start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"
        except Exception as e:
            report_title = f"{date_type or 'Date Range'}: {start_date} to {end_date}"
    else:
        report_title = "Current Product Prices"

    # PDF generation (keep your formatting)
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    try:
        logo_path = r"C:\Users\venki\Work\cms_1\static\cms_logo.png"
        print("Logo exists:", os.path.exists(logo_path))
        pdf.drawImage(logo_path, 1.5 * cm, height - 4 * cm, width=3 * cm, preserveAspectRatio=True)
    except Exception as e:
        print("Logo drawing failed:", e)

    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawCentredString(width / 2, height - 2.0 * cm, "Product Report")
    pdf.setFont("Helvetica", 12)
    pdf.drawCentredString(width / 2, height - 3.0 * cm, report_title)

    # Watermark
    pdf.saveState()
    pdf.setFont("Helvetica-Bold", 60)
    pdf.setFillColorRGB(0.85, 0.85, 0.85)
    pdf.translate(width / 2, height / 2)
    pdf.rotate(45)
    pdf.drawCentredString(0, 0, "CMS")
    pdf.restoreState()

    # Draw table
    MAX_ROWS_PER_PAGE = 40
    header = table_data[0]
    data_chunks = [table_data[i:i + MAX_ROWS_PER_PAGE] for i in range(1, len(table_data), MAX_ROWS_PER_PAGE)]

    top_margin = height - 5.5 * cm  # Start table just below title
    bottom_margin = 2 * cm

    for page_index, chunk in enumerate(data_chunks):
        if page_index > 0:
            pdf.showPage()

            # Title and watermark for additional pages
            pdf.setFont("Helvetica-Bold", 20)
            pdf.drawCentredString(width / 2, height - 2.0 * cm, "Product Report")
            pdf.setFont("Helvetica", 12)
            pdf.drawCentredString(width / 2, height - 3.0 * cm, report_title)

            pdf.saveState()
            pdf.setFont("Helvetica-Bold", 60)
            pdf.setFillColorRGB(0.85, 0.85, 0.85)
            pdf.translate(width / 2, height / 2)
            pdf.rotate(45)
            pdf.drawCentredString(0, 0, "CMS")
            pdf.restoreState()

        # Table rendering
        table_rows = [header] + chunk
        col_width = (width - 3 * cm) / len(header)
        table = Table(table_rows, colWidths=[col_width] * len(header))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d3d3d3")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5)
        ]))

        # Wrap with real page size
        table.wrapOn(pdf, 2 * cm, 5 * cm)
        table.drawOn(pdf, 1.5 * cm, height - 8 * cm - len(table_data) * 12)

    pdf.save()
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={brand}_Report.pdf"}
    )


# Save the view in History section
@router.post("/report/save")
def save_report_view(
    payload: SaveReportView,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    # Step 1: Save report name in client_reports
    new_report = CMSClientReport(
        userId=user_id,
        clientId=payload.client_id,
        name=payload.report_name,
        worksheetPrefix="",
        reportGraphTypeId=1,
        reportTypeId=1,
        preGenerate='excel',
        created=datetime.now(),
        modified=datetime.now()
    )
    db.add(new_report)
    db.commit()
    db.refresh(new_report)

    # Step 1: Resolve names to codes
    country_obj = db.query(CMSCountry).filter(CMSCountry.country == payload.country).first()
    category_obj = db.query(CMSCategory).filter(CMSCategory.sCategoryName == payload.category).first()
    subcategory_obj = db.query(CMSSubCategory).filter(CMSSubCategory.sCatSubName == payload.subcategory).first()
    brand_obj = db.query(CMSManufacturerBrand).filter(CMSManufacturerBrand.sBrandName == payload.brand).first()
    model_obj = db.query(CMSSummaryProductModel).filter(CMSSummaryProductModel.sProductModelNo == payload.model).first()
    retailer_obj = db.query(CrawlingWebsite).filter(CrawlingWebsite.companyName == payload.retailer).first()

    if not all([country_obj, category_obj, subcategory_obj, brand_obj, retailer_obj, model_obj]):
        raise HTTPException(status_code=404, detail="One or more values not found")
    
    # Create JSON with IDs
    filters_as_ids = {
        "country_code": country_obj.country_code,
        "category_code": category_obj.iCategoryCode,
        "subcategory_code": subcategory_obj.iCatSubCode,
        "brand_code": brand_obj.iManBrandCode,
        "model": model_obj.id,
        "retailer_id": retailer_obj.id,
        "model": model_obj.id,
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "date_type": payload.date_type
    }

    cache_entry = CMSCache(
        reportId=new_report.id,
        keyDetails="Saved filter metadata",  # placeholder string
        keyContents=json.dumps(filters_as_ids),
        data="{}",                           # empty object or actual data if you want to store it
        adCountData="{}",                   # placeholder
        fileGenerated="",                   # assume not generated yet
        lastCalculateDate=int(datetime.now().strftime("%Y%m%d")),
        deleted=0
    )

    db.add(cache_entry)

    # Step 4: Save filters into cms_client_reports_parameters using mst_parameter
    filter_map = {
        "country": country_obj.country_code,
        "category": category_obj.iCategoryCode,
        "subcategory": subcategory_obj.iCatSubCode,
        "brand": brand_obj.iManBrandCode,
        "model": model_obj.id,
        "retailer": retailer_obj.id,
        "model": model_obj.id,
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "date_type": payload.date_type
    }

    for key, value in filter_map.items():
        if value is not None:
            param = db.query(CMSMSTParameter).filter(CMSMSTParameter.code == key).first()
            if param:
                param_entry = CMSClientReportParameter(
                    reportId=new_report.id,
                    parameterId=param.id,
                    value=str(value)
                )
                db.add(param_entry)
    db.commit()

    return {"message": "Report view and filters saved", "report_id": new_report.id}

# Saved reports summary 
@router.get("/report-views", response_model=List[ReportSummary])
def get_report_views(
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    reports = db.query(CMSClientReport).filter(CMSClientReport.userId == user_id).order_by(CMSClientReport.created.desc()).all()
    return reports

# View action 
@router.get("/report/view/{report_id}")
def view_saved_report(report_id: int, db: Session = Depends(get_db), user_id: int = Depends(get_current_user_id)):
    # Step 1: Fetch the report
    report = db.query(CMSClientReport).filter(
        CMSClientReport.id == report_id, CMSClientReport.userId == user_id
    ).first()

    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    # Step 2: Get the latest cache data
    cache = (
        db.query(CMSCache)
        .filter(CMSCache.reportId == report_id)
        .order_by(CMSCache.id.desc())
        .first()
    )
    if not cache:
        raise HTTPException(status_code=404, detail="No cached data found for this report")

    filters = json.loads(cache.keyContents)

    # Step 3: Optionally, fetch readable names for display (optional)
    country = db.query(CMSCountry).filter(CMSCountry.country_code == filters.get("country_code")).first()
    category = db.query(CMSCategory).filter(CMSCategory.iCategoryCode == filters.get("category_code")).first()
    subcategory = db.query(CMSSubCategory).filter(CMSSubCategory.iCatSubCode == filters.get("subcategory_code"),CMSSubCategory.iCategoryCode == filters.get("category_code")).first()    
    brand = db.query(CMSManufacturerBrand).filter(CMSManufacturerBrand.iManBrandCode == filters.get("brand_code")).first()
    model = db.query(CMSSummaryProductModel).filter(CMSSummaryProductModel.id == filters.get("model")).first()
    retailer = db.query(CrawlingWebsite).filter(CrawlingWebsite.id == filters.get("retailer_id")).first()

    readable_filters = {
        "Country": country.country if country else "",
        "Category": category.sCategoryName if category else "",
        "Subcategory": subcategory.sCatSubName if subcategory else "",
        "Brand": brand.sBrandName if brand else "",
        "Retailer": retailer.companyName if retailer else "",
        "Model": model.sProductModelNo if model else "",
        "Start Date": filters.get("start_date"),
        "End Date": filters.get("end_date"),
        "Date Type": filters.get("date_type")
    }

    # Step 4: Return viewable data (or redirect to download endpoint)
    return JSONResponse({
        "report_id": report.id,
        "report_name": report.name,
        "filters": readable_filters,
        "data": cache.keyContents  
    })


# Get the saved report for update
@router.get("/report/edit/{report_id}")
def get_saved_report(report_id: int, db: Session = Depends(get_db)):
    # Step 1: Fetch the report
    report = db.query(CMSClientReport).filter(CMSClientReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    # Step 2: Get the associated cache entry (assuming 1-to-1 latest entry)
    cache = (
        db.query(CMSCache)
        .filter(CMSCache.reportId == report_id)
        .order_by(CMSCache.id.desc())  # just in case multiple entries exist
        .first()
    )

    if not cache:
        raise HTTPException(status_code=404, detail="Cache not found for this report")

    filters_json = json.loads(cache.keyContents)

    # Step 3: Resolve dropdown values from codes
    country = db.query(CMSCountry).filter(CMSCountry.country_code == filters_json.get("country_code")).first()
    category = db.query(CMSCategory).filter(CMSCategory.iCategoryCode == filters_json.get("category_code")).first()
    subcategory = db.query(CMSSubCategory).filter(CMSSubCategory.iCatSubCode == filters_json.get("subcategory_code"),CMSSubCategory.iCategoryCode == filters_json.get("category_code")).first()    
    brand = db.query(CMSManufacturerBrand).filter(CMSManufacturerBrand.iManBrandCode == filters_json.get("brand_code")).first()
    model = db.query(CMSSummaryProductModel).filter(CMSSummaryProductModel.id == filters_json.get("model")).first()
    retailer = db.query(CrawlingWebsite).filter(CrawlingWebsite.id == filters_json.get("retailer_id")).first()

    return {
        "report_id": report.id,
        "report_name": report.name,
        "filters": {
            "country": [{"id": country.country_code, "name": country.country}] if country else [],
            "category": [{"id": category.iCategoryCode, "name": category.sCategoryName}] if category else [],
            "subCategory": [{"id": subcategory.iCatSubCode, "name": subcategory.sCatSubName}] if subcategory else [],
            "retailer": [{"id": retailer.id, "name": retailer.companyName}] if retailer else [],
            "brand": [{"id": brand.iManBrandCode, "name": brand.sBrandName}] if brand else [],
            "model": [{"id": model.id, "name": model.sProductModelNo}] if model else [],
            "start_date": filters_json.get("start_date"),
            "end_date": filters_json.get("end_date"),
            "date_type": filters_json.get("date_type"),
        }
    }

# Updating the selected fields
@router.put("/report/update/{report_id}")
def update_report_view(
    report_id: int,
    payload: SaveReportView,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    # Step 1: Fetch and validate the report
    report = db.query(CMSClientReport).filter(
        CMSClientReport.id == report_id, CMSClientReport.userId == user_id
    ).first()

    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    # Step 2: Resolve names to codes
    country_obj = db.query(CMSCountry).filter(CMSCountry.country == payload.country).first()
    category_obj = db.query(CMSCategory).filter(CMSCategory.sCategoryName == payload.category).first()
    subcategory_obj = db.query(CMSSubCategory).filter(
        CMSSubCategory.sCatSubName == payload.subcategory,
        CMSSubCategory.iCategoryCode == category_obj.iCategoryCode
    ).first()
    brand_obj = db.query(CMSManufacturerBrand).filter(CMSManufacturerBrand.sBrandName == payload.brand).first()
    model_obj = db.query(CMSSummaryProductModel).filter(CMSSummaryProductModel.sProductModelNo == payload.model).first()
    retailer_obj = db.query(CrawlingWebsite).filter(CrawlingWebsite.companyName == payload.retailer).first()

    if not all([country_obj, category_obj, subcategory_obj, brand_obj, model_obj, retailer_obj]):
        raise HTTPException(status_code=404, detail="One or more values not found")

    # Step 3: Update `client_reports`
    report.name = payload.report_name
    report.modified = datetime.now()

    # Step 4: Update latest `cache` entry
    filters_as_ids = {
        "country_code": country_obj.country_code,
        "category_code": category_obj.iCategoryCode,
        "subcategory_code": subcategory_obj.iCatSubCode,
        "brand_code": brand_obj.iManBrandCode,
        "model": model_obj.id,
        "retailer_id": retailer_obj.id,
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "date_type": payload.date_type
    }

    latest_cache = (
        db.query(CMSCache)
        .filter(CMSCache.reportId == report_id)
        .order_by(CMSCache.id.desc())
        .first()
    )

    if latest_cache:
        latest_cache.keyContents = json.dumps(filters_as_ids)
        latest_cache.lastCalculateDate = int(datetime.now().strftime("%Y%m%d"))

    # Step 5: Clear old parameters and insert new ones
    db.query(CMSClientReportParameter).filter(CMSClientReportParameter.reportId == report_id).delete()

    filter_map = {
        "country": country_obj.country_code,
        "category": category_obj.iCategoryCode,
        "subcategory": subcategory_obj.iCatSubCode,
        "brand": brand_obj.iManBrandCode,
        "model": model_obj.id,
        "retailer": retailer_obj.id,
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "date_type": payload.date_type
    }

    for key, value in filter_map.items():
        if value is not None:
            param = db.query(CMSMSTParameter).filter(CMSMSTParameter.code == key).first()
            if param:
                param_entry = CMSClientReportParameter(
                    reportId=report.id,
                    parameterId=param.id,
                    value=str(value)
                )
                db.add(param_entry)

    db.commit()
    return {"message": "Report view updated successfully", "report_id": report.id}


# Delete the created reports
@router.delete("/reports/{report_id}", status_code=200)
def delete_report(
    report_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    # Step 1: Fetch the report
    report = db.query(CMSClientReport).filter(
        CMSClientReport.id == report_id,
        CMSClientReport.userId == user_id
    ).first()

    if not report:
        raise HTTPException(status_code=404, detail="Report not found or not authorized to delete")

    # Step 2: Delete related cache entries
    db.query(CMSCache).filter(CMSCache.reportId == report_id).delete()

    # Step 3: Delete related report parameters
    db.query(CMSClientReportParameter).filter(CMSClientReportParameter.reportId == report_id).delete()

    # Step 4: Delete the report
    db.delete(report)
    db.commit()

    return {"message": "Report and related cache data deleted successfully"}

    



